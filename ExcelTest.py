import openpyxl
import datetime,calendar
import requests
import json
import os

def getMonthDay(year,month):
    return calendar.monthrange(year,month)[1]

class excel:
    def __init__(self):
        self.__logData = None

    def createExcel(self,filename):
        __wb = openpyxl.Workbook()
        __wb.save(f"{filename}.xlsx")

    def saveExcel(self,**kwargs):
        self.__logData = self.getRequest(month=kwargs['month'],year=kwargs['year'])

        __filename = f"Absensi {calendar.month_name[kwargs['month']]} {kwargs['year']}"
        
        __maindir = os.getcwd()

        if __maindir.find('Absensi') < 0 :
            __maindir = os.path.join(__maindir,'Absensi')

        if not os.path.isfile(f'{__maindir}/data/{__filename}.xlsx'):
            self.createExcel(f'{__maindir}/data/{__filename}')

        __nameAttendance = {}
        for __item in self.__logData:
            __date = __item['date']
            for __att in __item['absensi']:
                __name = __att['nama']
                if __name not in __nameAttendance:
                    __nameAttendance[__name] = {'nama':__name,'absensi':[__date]}
                else:
                    __nameAttendance[__name]['absensi'].append(__date)

        self.__logData = list(__nameAttendance.values())
        __dayinmonth = getMonthDay(kwargs['year'],kwargs['month'])

        data = []
        data.append(tuple(['No','Nama']+[x for x in range(1,__dayinmonth+1)]))
        for i,log in enumerate(self.__logData):
            __logDate = [int(logDate.split('-')[0]) for logDate in log['absensi']]
            __absensi = [i+1,log['nama']]
            for date in range(1,__dayinmonth+1):
                __absensi.append('H') if date in(__logDate) else __absensi.append(' ')
            data.append(tuple(__absensi))
            
        __wb = openpyxl.load_workbook(f'{__maindir}/data/{__filename}.xlsx',data_only=True)
        __sheet = __wb.active
        _maxCol = openpyxl.utils.cell.get_column_letter(len(data[0]))

        for i, value in enumerate(data):
            for row in __sheet[f'A{i+1}:{_maxCol}{i+1}']:
                for j, cell in enumerate(row):
                    cell.value = value[j]
                    
        '''
        for row in __sheet[f'A1:{_maxCol}{_maxRow}']:
            for cell in row:
                cell.value = None
        '''
        
        __wb.save(f'{__maindir}/data/{__filename}.xlsx')

    def getRequest(self, **kwargs):
        __params = params = {
                'bulan':kwargs['month'],
                'tahun':kwargs['year']
            }
        __headers = {
                'User-Agent': 'Custom'
            }
        __r = requests.get('http://robotik.pkm.unp.ac.id/api/absensi/get/',params=__params,headers=__headers)

        return __r.json()

excel().saveExcel(month=4,year=2023)